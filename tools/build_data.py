#!/usr/bin/env python3
"""
Converts case study Excel workbooks (excel/<sector>/*.xlsx) into the .js
data files the website reads (data/<sector>/*.js).

Usage:
    python3 tools/build_data.py

Run it every time you edit an Excel file, then commit + push the
regenerated data/*.js files (git add -A && git commit -m "..." && git push).

Requires: openpyxl (one-time setup: python3 -m pip install --user openpyxl)
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Missing dependency 'openpyxl'. Install it once with:")
    print("    python3 -m pip install --user openpyxl")
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent
EXCEL_DIR = ROOT / "excel"
DATA_DIR = ROOT / "data"


def sheet_rows(ws):
    """Yields each data row (after the header row) as a list of cell values."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return
    for row in rows[1:]:
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue
        yield row


def as_str(value):
    return "" if value is None else str(value).strip()


def as_num(value):
    if value is None or str(value).strip() == "":
        return 0
    return float(value) if not float(value).is_integer() else int(value)


def read_info(wb):
    ws = wb["Info"]
    info = {}
    for row in sheet_rows(ws):
        field, value = row[0], row[1]
        if field is None:
            continue
        info[as_str(field)] = value
    return info


def read_pairs(wb, sheet_name, key_a, key_b):
    """Two-column sheet -> list of {key_a: ..., key_b: ...} dicts."""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    return [
        {key_a: as_str(row[0]), key_b: as_str(row[1])}
        for row in sheet_rows(ws)
    ]


def build_project(xlsx_path: Path, sector_folder: str):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    info = read_info(wb)

    project = {
        "id": as_str(info.get("id")),
        "sector": as_str(info.get("sector")),
        "sectorLabel": as_str(info.get("sectorLabel")),
        "status": as_str(info.get("status")) or "example",
        "name": as_str(info.get("name")),
        "location": as_str(info.get("location")),
        "size": as_str(info.get("size")),
        "tagline": as_str(info.get("tagline")),
    }

    if not project["id"]:
        raise ValueError(f"'Info' sheet is missing a value for 'id' in {xlsx_path.name}")

    project["metrics"] = read_pairs(wb, "Metrics", "label", "value")
    project["facts"] = read_pairs(wb, "Facts", "label", "value")

    project["executiveSummary"] = [
        as_str(row[0]) for row in sheet_rows(wb["ExecutiveSummary"]) if as_str(row[0])
    ]

    scenario_rows = []
    for row in sheet_rows(wb["ScenarioTable"]):
        scenario_rows.append({
            "label": as_str(row[0]),
            "unit": as_str(row[1]),
            "p50": as_num(row[2]),
            "p75": as_num(row[3]),
            "p90": as_num(row[4]),
            "p99": as_num(row[5]),
        })
    project["scenarioTable"] = {
        "caption": as_str(info.get("scenarioCaption")) or "Results by probabilistic scenario",
        "rows": scenario_rows,
    }

    project["callouts"] = read_pairs(wb, "Callouts", "value", "text")

    sensitivities = []
    for row in sheet_rows(wb["Sensitivities"]):
        sensitivities.append({
            "label": as_str(row[0]),
            "low": as_num(row[1]),
            "high": as_num(row[2]),
            "note": as_str(row[3]),
        })
    project["sensitivities"] = sensitivities

    debt_points = []
    for row in sheet_rows(wb["DebtProfile"]):
        debt_points.append({"year": as_num(row[0]), "dscr": as_num(row[1])})
    project["debtProfile"] = {
        "covenantMin": as_num(info.get("covenantMin")),
        "points": debt_points,
    }

    project["insightsPE"] = read_pairs(wb, "InsightsPE", "title", "body")
    project["insightsLenders"] = read_pairs(wb, "InsightsLenders", "title", "body")

    model_name = as_str(info.get("modelFileName"))
    if model_name:
        project["modelFile"] = {
            "name": model_name,
            "meta": as_str(info.get("modelFileMeta")),
            "href": as_str(info.get("modelFileHref")) or "#",
        }
    else:
        project["modelFile"] = None

    assumptions = {"revenue": [], "capex": [], "opex": [], "debt": [], "equity": []}
    for row in sheet_rows(wb["Assumptions"]):
        category = as_str(row[0]).lower()
        if category not in assumptions:
            print(f"  Warning: unknown assumption category '{row[0]}' in {xlsx_path.name} (skipped)")
            continue
        assumptions[category].append({
            "parameter": as_str(row[1]),
            "value": as_str(row[2]),
            "source": as_str(row[3]),
        })
    project["assumptions"] = assumptions

    sources = []
    for row in sheet_rows(wb["Sources"]):
        sources.append({
            "org": as_str(row[0]),
            "title": as_str(row[1]),
            "desc": as_str(row[2]),
            "href": as_str(row[3]) or "#",
        })
    project["sources"] = sources

    return project


def write_js(project: dict, out_path: Path):
    body = json.dumps(project, indent=2, ensure_ascii=False)
    js = (
        "window.CASE_STUDIES = window.CASE_STUDIES || [];\n\n"
        f"window.CASE_STUDIES.push({body});\n"
    )
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(js, encoding="utf-8")


def main():
    if not EXCEL_DIR.exists():
        print(f"No excel/ directory found at {EXCEL_DIR}")
        sys.exit(1)

    built = 0
    for xlsx_path in sorted(EXCEL_DIR.rglob("*.xlsx")):
        if xlsx_path.name.startswith("_") or xlsx_path.name.startswith("~$"):
            continue  # skip templates and Excel lock files

        sector_folder = xlsx_path.parent.name
        try:
            project = build_project(xlsx_path, sector_folder)
        except Exception as e:
            print(f"FAILED  {xlsx_path.relative_to(ROOT)}: {e}")
            continue

        out_path = DATA_DIR / sector_folder / f"{xlsx_path.stem}.js"
        write_js(project, out_path)
        print(f"OK      {xlsx_path.relative_to(ROOT)} -> {out_path.relative_to(ROOT)}")
        built += 1

    print(f"\n{built} case study file(s) generated.")


if __name__ == "__main__":
    main()
